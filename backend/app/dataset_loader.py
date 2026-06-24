from __future__ import annotations

from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATASET_DIR = PROJECT_ROOT / "datasets"

DATA_SOURCE = {
    "name": "SkillsFuture Jobs-Skills Dataset",
    "release": "Q2 2026",
    "note": "Uses the official SkillsFuture role, TSC mapping, and unique skill vocabulary Excel files.",
}

FRAMEWORK_FILE = DATASET_DIR / "jobsandskills-skillsfuture-skills-framework-dataset.xlsx"
MAPPING_FILE = DATASET_DIR / "jobsandskills-skillsfuture-tsc-to-unique-skills-mapping.xlsx"
UNIQUE_SKILLS_FILE = DATASET_DIR / "jobsandskills-skillsfuture-unique-skills-list.xlsx"


@lru_cache(maxsize=1)
def dataset_summary() -> dict:
    framework_roles, role_skill_links = read_framework_data()
    mapping = read_mapping()
    skill_descriptions = read_unique_skill_descriptions()
    unique_skills = set(skill_descriptions)

    normalised_pairs = 0
    mapping_by_code = mapping["by_code"]
    mapping_by_title = mapping["by_title"]

    for link in role_skill_links:
        unique_skill = mapping_by_code.get(link["code"]) or mapping_by_title.get(link["title"])
        if unique_skill and unique_skill in unique_skills:
            normalised_pairs += 1

    return {
        "uses_all_three_files": True,
        "files": {
            "framework": FRAMEWORK_FILE.name,
            "tsc_to_unique_mapping": MAPPING_FILE.name,
            "unique_skills": UNIQUE_SKILLS_FILE.name,
        },
        "framework_roles": len(framework_roles),
        "role_skill_links": len(role_skill_links),
        "mapped_unique_skills": len(set(mapping_by_code.values()) | set(mapping_by_title.values())),
        "unique_skill_vocabulary": len(unique_skills),
        "normalised_role_skill_pairs": normalised_pairs,
        "sample_roles": sorted(framework_roles)[:8],
    }


@lru_cache(maxsize=1)
def official_role_profiles() -> dict[str, dict]:
    framework_roles, role_skill_links = read_framework_data()
    mapping = read_mapping()
    skill_descriptions = read_unique_skill_descriptions()
    unique_skills = set(skill_descriptions)
    by_code = mapping["by_code"]
    by_title = mapping["by_title"]
    profiles: dict[str, dict] = {
        title: {
            "id": slugify(title),
            "title": title,
            "sector": role["sector"],
            "track": role["track"],
            "description": role["description"],
            "skills": [],
            "skill_descriptions": {},
            "required": {},
        }
        for title, role in framework_roles.items()
    }

    for link in role_skill_links:
        profile = profiles.get(link["role"])
        if not profile:
            continue

        unique_skill = by_code.get(link["code"]) or by_title.get(link["title"]) or link["title"]
        if unique_skill not in unique_skills:
            continue

        if unique_skill not in profile["skills"]:
            profile["skills"].append(unique_skill)

        if unique_skill in skill_descriptions:
            profile["skill_descriptions"][unique_skill] = skill_descriptions[unique_skill]

        weight = proficiency_weight(link["proficiency"])
        profile["required"][unique_skill] = max(profile["required"].get(unique_skill, 0), weight)

    return {
        title: profile
        for title, profile in profiles.items()
        if profile["skills"]
    }


def search_official_roles(query: str = "", limit: int = 200) -> list[dict]:
    query_normalized = query.casefold().strip()
    profiles = official_role_profiles().values()

    if query_normalized:
        profiles = [
            profile
            for profile in profiles
            if query_normalized in profile["title"].casefold()
            or query_normalized in profile["sector"].casefold()
            or query_normalized in profile["track"].casefold()
        ]

    safe_limit = max(1, min(limit, 2500))

    return sorted(
        [
            {
                "id": profile["id"],
                "title": profile["title"],
                "sector": profile["sector"],
                "track": profile["track"],
                "skills": profile["skills"],
                "skill_count": len(profile["skills"]),
            }
            for profile in profiles
        ],
        key=lambda profile: (profile["sector"], profile["title"]),
    )[:safe_limit]


def get_official_role(title_or_id: str) -> dict | None:
    profiles = official_role_profiles()
    if title_or_id in profiles:
        return profiles[title_or_id]

    normalized = title_or_id.casefold().strip()
    for profile in profiles.values():
        if profile["id"] == title_or_id or profile["title"].casefold() == normalized:
            return profile

    return None


def read_framework_data() -> tuple[dict[str, dict], list[dict[str, str]]]:
    workbook = load_workbook(FRAMEWORK_FILE, read_only=True, data_only=True)
    roles_sheet = workbook["Job Role_Description"]
    links_sheet = workbook["Job Role_TCS_CCS"]

    roles = {}
    for row in roles_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue

        role_title = clean(row[2])
        roles[role_title] = {
            "sector": clean(row[0]),
            "track": clean(row[1]),
            "description": clean(row[3]),
            "performance_expectation": clean(row[4]),
        }

    links = []
    for row in links_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[2] or not row[3]:
            continue

        links.append(
            {
                "sector": clean(row[0]),
                "track": clean(row[1]),
                "role": clean(row[2]),
                "title": clean(row[3]),
                "type": clean(row[4]),
                "proficiency": clean(row[5]),
                "code": clean(row[6]),
            }
        )

    workbook.close()
    return roles, links


def read_mapping() -> dict[str, dict[str, str]]:
    workbook = load_workbook(MAPPING_FILE, read_only=True, data_only=True)
    sheet = workbook["data"]
    by_code: dict[str, str] = {}
    by_title: dict[str, str] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        code = clean(row[0])
        framework_title = clean(row[1])
        updated_title = clean(row[10])

        if code and updated_title:
            by_code[code] = updated_title
        if framework_title and updated_title:
            by_title[framework_title] = updated_title

    workbook.close()
    return {"by_code": by_code, "by_title": by_title}


def read_unique_skills() -> set[str]:
    return set(read_unique_skill_descriptions())


@lru_cache(maxsize=1)
def read_unique_skill_descriptions() -> dict[str, str]:
    workbook = load_workbook(UNIQUE_SKILLS_FILE, read_only=True, data_only=True)
    sheet = workbook["Unique Skills List"]
    skills = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        title = clean(row[0])
        description = clean(row[1])
        skills[title] = description

    workbook.close()
    return skills


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def proficiency_weight(value: str) -> int:
    try:
        level = int(float(value))
    except (TypeError, ValueError):
        level = 3

    return max(45, min(96, 42 + level * 9))


def slugify(value: str) -> str:
    output = []
    previous_dash = False

    for character in value.lower():
        if character.isalnum():
            output.append(character)
            previous_dash = False
        elif not previous_dash:
            output.append("-")
            previous_dash = True

    return "".join(output).strip("-")
